#Send emails based on payment status in spreadsheet.
import openpyxl, smtplib, sys
#Open spreadsheet and get the lastest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
lastestMonth = sheet.cell(row = 1, column = lastCol).value
#TODO: Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
	payment = sheet.cell(row = r, column = lastCol).value
	if payment != 'paid':
		name = sheet.cell(row = r, column = 1).value
		email = sheet.cell(row = r, column = 2).value
		unpaidMembers[name] = email
#TODO: Login to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
myemail = '17it142@sict.udn.vn'
smtpObj.login(myemail, sys.argv[1])
print('Login successed!')
#TODO: Send out reminder emails.
for name, email in unpaidMembers.items():
	body = "Subject %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (lastestMonth, name, lastestMonth)
	print('Sending email to %s...' % email)
	sendmailStatus = smtpObj.sendmail(myemail, email, body)
	if sendmailStatus != {}:
		print('There was a problem sending mail to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
