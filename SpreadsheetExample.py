import openpyxl
#from openpyxl.cell import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')
#print(type(wb))
print(wb.sheetnames) #List tất cả tên bảng
sheet = wb['Sheet1'] #Chọn bảng
another = wb.active
"""
#Lấy dữ liệu từ 1 ô
cell = sheet['A1']
row = c.row #1
col = c.column #A
coordinate = c.cordinate #A1
print(sheet['A2'].value)

print(sheet.max_row)
print(sheet.max_column)

get_column_letter(900)
=> AHP

print(sheet.max_row) #7
print(sheet.max_column) #3

for i in range(1, sheet.max_row + 1):
	for j in range(1, sheet.max_column + 1):
		print(sheet.cell(column = j, row = i).value, end= '\t\t\t\t')
	print()

Quá trình xử lý:
1. Nạp module openpyxl
2. Gọi phương thức openpyxl.load_work_book()
3. Lấy đối tượng Workbook ở trên
4. Đọc bảng active hoặc lấy theo tên bảng
5. Lấy đối tượng bảng tính ở trên
6. Sử dụng chỉ số hoặc phương thức cell(row, column)
7. Lấy đối tượng Cell ở trên
8. Đọc giá trị bằng Cell.value

--- WRITING EXCEL DOCUMENTS ---
wb = openpyxl.Workbook() #Create a new Workbook
wb.create_sheet() #Automatic create a new sheet_.
wb.create_sheet(index = i, title = '')
wb.remove(worksheet) or del wb[sheetname] #Delete a sheet
wb.save('') #Save

--- SETTING THE FONT STYLE ---
newFont = Font(name = '', size = i, italic/bold = True/False)
sheet['A1'].font = newFont

--- FORMULAS ---
sheet['A3'] = '=SUM(A1:A2)'

--- ADJUSTING ROWS AND COLUMNS ---
sheet.row_dimensions[1].height = h(0 -> 409, default is 12.75)
sheet.row_dimensions['A'].width = w(0 -> 255, default is 8.43 character)
height/row = 0 -> Hidden

--- MERGING AND UNMERGING CELLS ---
sheet.merge_cells('A1:Z5')

--- FREEZE PANES ---
sheet.freeze_panes = cell #Đóng băng tất cả hàng ở trên và cột ở bên trái cell.
sheet.freeze_panes = 'A1' or sheet.freeze_panes = None

--- CHARTS ---
To make a chart, you need to do the following:
1. Create a Reference object from a rectangular selection of cells.
-> openpyxl.chart.Reference(x, x1, y1, x2, y2)
	+ x is Worksheet object containing chart data.
	+ x1, x2 is col
	+ y1, y2 is row
2. Create a Series object by passing in the Reference object.
3. Create a Chart object.
4. Append the Series object to the Chart object.
5. Add the Chart object to the Worksheet object, optionally
specifying which cell the top left corner of the chart should be positioned.
-> Example:
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
	sheet['A' + str(i)] = i
refObj = openpyxl.chart.Refercence(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()/LineChart()/ScatterChart()/PieChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
"""
for rowOfCellObj in sheet['B1': 'B7']:
	for cellObj in rowOfCellObj:
		print('\n',cellObj.coordinate, cellObj.value, '\n')
	print('--- END OF ROW ---')