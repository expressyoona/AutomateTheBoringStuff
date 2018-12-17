import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb.active
fontBold = Font(bold = True)
#Init
for i in range(2, 11):
	sheet['A' + str(i)] = i - 1
	sheet['A' + str(i)].font = fontBold
	sheet.cell(column = i, row = 1).value = i - 1
	sheet.cell(column = i, row = 1).font = fontBold
for row in sheet['B2:J10']:
	for cell in row:
		#B2 = B1 * A2
		r = 'A' + str(cell.row)
		c = '$' + str(cell.column) + '$' + str(1)
		cell.value = '=' + r + '*' + c
		#print(cell.coordinate, ' = ', r, ' * ',c)
		#print(cell.coordinate, ' -> ',cell.row, ', ', cell.column)
wb.save('multiplication.xlsx')
print('Saved!')
