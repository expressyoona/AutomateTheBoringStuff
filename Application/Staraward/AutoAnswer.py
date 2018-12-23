import pyautogui
import time
import openpyxl
import random


bt_1 = (1199, 247)
continue_bt = (685, 362)
user_field = (500, 356)
exit_bt = (1300, 173)

#Default password is sictudnvn
pw = 'sictudnvn'

wb = openpyxl.load_workbook('Register.xlsx')

sheet = wb.active

row = sheet.max_row

for r in range(2, row):

	pyautogui.click(user_field[0], user_field[1])
	username = sheet.cell(row = r, column = 1).value
	
	pyautogui.typewrite(username + '\t' + pw + '\t')
	pyautogui.typewrite(['enter'])
	time.sleep(5)
	pyautogui.click(bt_1[0], bt_1[1])
	
	#Join contest
	#Go to first question
	pyautogui.typewrite(['tab'])
	time.sleep(2)
	pyautogui.typewrite(['tab'])
	time.sleep(2)
	pyautogui.typewrite(['tab'])
	time.sleep(2)
	#Answer 30 questions
	for i in range(0, 30):
		#Select answer A
		pyautogui.typewrite(['tab', 'space'])
		"""
		#Randomize answer
		answer = ['down'] * random.randint(0, 3)
		if answer:
			pyautogui.typewrite(answer)
		"""
		#Sleep one second
		time.sleep(1)
	
	#Submit
	pyautogui.typewrite(['enter'])
	time.sleep(2)

	pyautogui.typewrite(['enter'])
	time.sleep(3)

	pyautogui.click(continue_bt[0], continue_bt[1])
	time.sleep(3)

print('Done!')