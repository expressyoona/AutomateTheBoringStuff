#Convert name
def convert(name):
	name = name.lower().split(' ')
	rs = ''
	for i in range(0, len(name) - 1):
		rs += name[i][0]
	return rs + name[len(name) - 1]

import openpyxl
#Source file is Book.xlsx
wb = openpyxl.load_workbook('Book.xlsx')

sheet = wb['Main']

rows = sheet.max_row
cols = sheet.max_column
#Cols = 2

for i in range(1, rows + 1):
	name = sheet.cell(column = 2, row = i).value
	sheet['C' + str(i)] = convert(name) + '.17it1@sict.udn.vn'
  
wb.save('Mail.xlsx')

print('Done!')
