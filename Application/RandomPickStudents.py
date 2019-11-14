import openpyxl
import random
from unidecode import unidecode

SOURCE_FILENAME = 'DanhSachLop_Info.xlsx'
DESTINATION_FILENAME = '17IT1_NCKH.xlsx'
# MAX = 37
N = 12

wb = openpyxl.load_workbook(SOURCE_FILENAME)
sheet = wb.active

def getEmail(fullname):
	fullname = unidecode(fullname.lower())
	return "".join(w[0] for w in fullname.split(' ')[:-1]) + fullname.split(' ')[-1] + '.' + '17it1@sict.udn.vn'


new_workbook = openpyxl.Workbook()
ws = new_workbook.create_sheet(title='17IT1')


L = random.sample(list(range(1, MAX + 1)), N)
for index, element in enumerate(L):
	ws['A' + str(index + 1)] = index + 1
	#Get ID
	ws['B' + str(index + 1)] = sheet.cell(column = 3, row=element + 2).value
	#Get name
	fullname = sheet.cell(column = 4, row=element + 2).value
	ws['C' + str(index + 1)] = fullname
	ws['D' + str(index + 1)] = getEmail(fullname)

new_workbook.save(DESTINATION_FILENAME)
print('Done!')
